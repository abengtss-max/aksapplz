"""Add a 'Reference / Information' column to the 'Accelerator - Terraform' tab
of the Azure Landing Zone accelerator checklist, with a Microsoft Learn link
per row. Style mirrors the AKS landing-zone checklist (blue 'Microsoft Learn'
hyperlink). Does not touch any other tab. Makes a .bak backup first.
"""
import shutil
import sys
from copy import copy

import openpyxl
from openpyxl.styles import Font

SRC = sys.argv[1]
TAB = "Accelerator - Terraform"
HEADER_ROW = 3
HEADER_TEXT = "Reference / Information"
LINK_TEXT = "Microsoft Learn \u2197"

# row number -> URL
LINKS = {
    # --- Scenario section (network topology choices) ---
    6:  "https://learn.microsoft.com/azure/cloud-adoption-framework/ready/azure-best-practices/define-an-azure-network-topology",  # Multi-Region Hub & Spoke + Azure Firewall
    7:  "https://learn.microsoft.com/azure/cloud-adoption-framework/ready/azure-best-practices/virtual-wan-network-topology",       # Multi-Region Virtual WAN + Azure Firewall
    8:  "https://learn.microsoft.com/azure/architecture/networking/guide/network-virtual-appliance",                                # Multi-Region Hub & Spoke + NVA
    9:  "https://learn.microsoft.com/azure/virtual-wan/about-nva-hub",                                                              # Multi-Region Virtual WAN + NVA
    10: "https://learn.microsoft.com/azure/cloud-adoption-framework/ready/enterprise-scale/management-group-and-subscription-organization",  # Mgmt Groups / Policy only
    11: "https://learn.microsoft.com/azure/architecture/networking/architecture/hub-spoke",                                         # Single-Region Hub & Spoke + Azure Firewall
    12: "https://learn.microsoft.com/azure/cloud-adoption-framework/ready/azure-best-practices/virtual-wan-network-topology",       # Single-Region Virtual WAN + Azure Firewall
    13: "https://learn.microsoft.com/azure/architecture/networking/guide/network-virtual-appliance",                                # Single-Region Hub & Spoke + NVA
    14: "https://learn.microsoft.com/azure/virtual-wan/about-nva-hub",                                                              # Single-Region Virtual WAN + NVA
    # --- Options section ---
    17: "https://learn.microsoft.com/azure/cloud-adoption-framework/ready/azure-best-practices/resource-naming",                    # Resource naming convention
    18: "https://learn.microsoft.com/azure/governance/management-groups/overview",                                                  # Custom management group names
    19: "https://learn.microsoft.com/azure/ddos-protection/ddos-protection-overview",                                               # DDoS Protection Plan
    20: "https://learn.microsoft.com/azure/cloud-adoption-framework/ready/azure-best-practices/private-link-and-dns-integration-at-scale",  # Private DNS
    21: "https://learn.microsoft.com/azure/bastion/bastion-overview",                                                               # Bastion Host
    22: "https://learn.microsoft.com/azure/vpn-gateway/vpn-gateway-about-vpngateways",                                              # VPN Gateway
    23: "https://learn.microsoft.com/azure/expressroute/expressroute-about-virtual-network-gateways",                              # ExpressRoute Gateway
    24: "https://learn.microsoft.com/azure/reliability/regions-overview",                                                           # Deploy to more than 2 regions
    25: "https://learn.microsoft.com/azure/cloud-adoption-framework/ready/azure-best-practices/plan-for-ip-addressing",             # IP Addressing
    26: "https://learn.microsoft.com/azure/governance/policy/concepts/assignment-structure#enforcement-mode",                       # Change policy enforcement mode
    27: "https://learn.microsoft.com/azure/governance/policy/concepts/assignment-structure",                                        # Remove a policy assignment
    28: "https://learn.microsoft.com/azure/azure-monitor/agents/azure-monitor-agent-overview",                                      # Turn off Azure Monitoring Agent
    29: "https://azure.github.io/azure-monitor-baseline-alerts/welcome/",                                                           # Azure Monitor Baseline Alerts (AMBA)
    30: "https://learn.microsoft.com/azure/defender-for-cloud/defender-for-cloud-introduction",                                     # Turn off Defender Plans
    31: "https://learn.microsoft.com/security/zero-trust/zero-trust-overview",                                                      # Zero Trust Security
    32: "https://learn.microsoft.com/industry/sovereignty/sovereign-landing-zone",                                                  # Sovereign Landing Zone
}

# backup
bak = SRC + ".bak"
shutil.copyfile(SRC, bak)
print("Backup written:", bak)

wb = openpyxl.load_workbook(SRC)
ws = wb[TAB]

new_col = ws.max_column + 1  # append to the right of existing columns
col_letter = ws.cell(row=1, column=new_col).column_letter

# Header cell: copy style from the existing last header cell (e.g. 'Example')
src_hdr = ws.cell(row=HEADER_ROW, column=ws.max_column)
hdr = ws.cell(row=HEADER_ROW, column=new_col, value=HEADER_TEXT)
if src_hdr.has_style:
    hdr.font = copy(src_hdr.font)
    hdr.fill = copy(src_hdr.fill)
    hdr.border = copy(src_hdr.border)
    hdr.alignment = copy(src_hdr.alignment)
    hdr.number_format = src_hdr.number_format

link_font = Font(name=hdr.font.name or "Calibri", size=hdr.font.size or 11,
                 color="0563C1", underline="single")

for row, url in LINKS.items():
    # copy border/fill from the neighbouring (left) cell so the grid matches
    left = ws.cell(row=row, column=new_col - 1)
    c = ws.cell(row=row, column=new_col, value=LINK_TEXT)
    c.hyperlink = url
    c.font = link_font
    if left.has_style:
        c.border = copy(left.border)
        c.fill = copy(left.fill)
        c.alignment = copy(left.alignment)

# match column width to the 'Example' column (or a sensible default)
src_dim = ws.column_dimensions.get(src_hdr.column_letter)
ws.column_dimensions[col_letter].width = (src_dim.width if src_dim and src_dim.width else 24)

wb.save(SRC)
print(f"Added column {col_letter} '{HEADER_TEXT}' with {len(LINKS)} links to '{TAB}'.")
