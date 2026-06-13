"""
Generate the AKS Application Landing Zone planning checklist (checklist.xlsx).

The workbook is the planning artifact for `Deploy-AKSLandingZone`. Each yellow
"Your Value" cell in Tab 1 maps 1:1 to a prompt in the wizard and to a setting
in config/inputs.yaml.

Regenerate with:
    pip install openpyxl
    python config/generate_checklist.py
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ─── Styles ────────────────────────────────────────────────────────────────
header_font   = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
header_fill   = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
section_font  = Font(name="Segoe UI", size=11, bold=True, color="0078D4")
section_fill  = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
normal_font   = Font(name="Segoe UI", size=10)
input_fill    = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
title_font    = Font(name="Segoe UI", size=14, bold=True, color="0078D4")
subtitle_font = Font(name="Segoe UI", size=10, italic=True, color="555555")
link_font     = Font(name="Segoe UI", size=10, color="0563C1", underline="single")

thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
wrap_alignment   = Alignment(wrap_text=True, vertical="top")
center_alignment = Alignment(horizontal="center", vertical="center")


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def style_section(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border


def style_data(ws, row, cols, input_col=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.alignment = wrap_alignment
        cell.border = thin_border
        if input_col and c == input_col:
            cell.fill = input_fill


def dropdown(ws, row, col, values):
    dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=col))


# ─── Microsoft Learn / GitHub Docs reference per setting ─────────────────────
REFERENCES = {
    # Scenario & topology
    "scenario": "https://learn.microsoft.com/azure/architecture/reference-architectures/containers/aks/baseline-aks",
    "secondary_location": "https://learn.microsoft.com/azure/aks/ha-dr-overview",
    "topology": "https://learn.microsoft.com/azure/architecture/networking/architecture/hub-spoke",
    "global_lb_type": "https://learn.microsoft.com/azure/architecture/guide/technology-choices/load-balancing-overview",
    # Where to deploy
    "bootstrap_location": "https://learn.microsoft.com/azure/reliability/regions-overview",
    "aks_landing_zone_subscription_id": "https://learn.microsoft.com/azure/cloud-adoption-framework/ready/landing-zone/design-area/resource-org-subscriptions",
    "connectivity_subscription_id": "https://learn.microsoft.com/azure/cloud-adoption-framework/ready/landing-zone/design-area/network-topology-and-connectivity",
    # Hub network
    "hub_vnet_resource_id": "https://learn.microsoft.com/azure/architecture/networking/architecture/hub-spoke",
    "hub_firewall_private_ip": "https://learn.microsoft.com/azure/aks/limit-egress-traffic",
    "hub_vnet_address_space": "https://learn.microsoft.com/azure/virtual-network/virtual-networks-overview",
    "hub_firewall_subnet_address_prefix": "https://learn.microsoft.com/azure/firewall/firewall-faq",
    "hub_deploy_firewall": "https://learn.microsoft.com/azure/firewall/overview",
    "hub_firewall_sku_tier": "https://learn.microsoft.com/azure/firewall/choose-firewall-sku",
    # Spoke subnets
    "spoke_vnet_address_space": "https://learn.microsoft.com/azure/aks/concepts-network",
    "subnet_address_prefix_aks_system_nodes": "https://learn.microsoft.com/azure/aks/azure-cni-overlay",
    "subnet_address_prefix_aks_user_nodes": "https://learn.microsoft.com/azure/aks/azure-cni-overlay",
    "subnet_address_prefix_aks_api_server": "https://learn.microsoft.com/azure/aks/api-server-vnet-integration",
    "subnet_address_prefix_app_gateway": "https://learn.microsoft.com/azure/application-gateway/ingress-controller-overview",
    "subnet_address_prefix_private_endpoints": "https://learn.microsoft.com/azure/private-link/private-endpoint-overview",
    "subnet_address_prefix_ingress": "https://learn.microsoft.com/azure/aks/internal-lb",
    "subnet_address_prefix_agc": "https://learn.microsoft.com/azure/application-gateway/for-containers/overview",
    # Cluster settings
    "kubernetes_version": "https://learn.microsoft.com/azure/aks/supported-kubernetes-versions",
    "aks_sku_tier": "https://learn.microsoft.com/azure/aks/free-standard-pricing-tiers",
    "aks_private_cluster": "https://learn.microsoft.com/azure/aks/private-clusters",
    "aks_admin_group_object_ids": "https://learn.microsoft.com/azure/aks/azure-ad-rbac",
    # State & naming
    "bootstrap_subscription_id": "https://learn.microsoft.com/azure/developer/terraform/store-state-in-azure-storage",
    "service_name": "https://learn.microsoft.com/azure/cloud-adoption-framework/ready/azure-best-practices/resource-naming",
    "environment_name": "https://learn.microsoft.com/azure/cloud-adoption-framework/ready/azure-best-practices/resource-naming",
    "postfix_number": "https://learn.microsoft.com/azure/cloud-adoption-framework/ready/azure-best-practices/resource-naming",
    # Runners & GitHub
    "use_self_hosted_runners": "https://docs.github.com/actions/hosting-your-own-runners/managing-self-hosted-runners/about-self-hosted-runners",
    "use_private_networking": "https://learn.microsoft.com/azure/container-registry/container-registry-private-link",
    "github_personal_access_token": "https://docs.github.com/authentication/keeping-your-account-and-data-secure/managing-your-personal-access-tokens",
    "github_runners_personal_access_token": "https://docs.github.com/authentication/keeping-your-account-and-data-secure/managing-your-personal-access-tokens",
    "github_organization_name": "https://docs.github.com/organizations/collaborating-with-groups-in-organizations/about-organizations",
    "apply_approvers": "https://docs.github.com/actions/deployment/targeting-different-environments/using-environments-for-deployment",
    # Add-ons
    "enable_defender": "https://learn.microsoft.com/azure/defender-for-cloud/defender-for-containers-introduction",
    "enable_workload_identity": "https://learn.microsoft.com/azure/aks/workload-identity-overview",
    "enable_azure_policy": "https://learn.microsoft.com/azure/aks/use-azure-policy",
    "enable_prometheus": "https://learn.microsoft.com/azure/azure-monitor/essentials/prometheus-metrics-overview",
    "enable_grafana": "https://learn.microsoft.com/azure/managed-grafana/overview",
    "enable_app_gateway": "https://learn.microsoft.com/azure/application-gateway/ingress-controller-overview",
    "enable_agc": "https://learn.microsoft.com/azure/application-gateway/for-containers/overview",
    "enable_keda": "https://learn.microsoft.com/azure/aks/keda-about",
    "enable_vpa": "https://learn.microsoft.com/azure/aks/vertical-pod-autoscaler",
    "enable_node_auto_provisioning": "https://learn.microsoft.com/azure/aks/node-autoprovision",
    "enable_istio": "https://learn.microsoft.com/azure/aks/istio-about",
    "enable_flux": "https://learn.microsoft.com/azure/azure-arc/kubernetes/conceptual-gitops-flux2",
    "enable_dapr": "https://learn.microsoft.com/azure/aks/dapr",
    "enable_fips": "https://learn.microsoft.com/azure/aks/enable-fips-nodes",
    "enable_backup": "https://learn.microsoft.com/azure/backup/azure-kubernetes-service-backup-overview",
    "enable_cost_analysis": "https://learn.microsoft.com/azure/aks/cost-analysis",
    # Advanced cluster settings (Tab 2)
    "system_node_pool.vm_size": "https://learn.microsoft.com/azure/aks/use-system-pools",
    "system_node_pool.min_count": "https://learn.microsoft.com/azure/aks/cluster-autoscaler-overview",
    "system_node_pool.max_count": "https://learn.microsoft.com/azure/aks/cluster-autoscaler-overview",
    "user_node_pool.vm_size": "https://learn.microsoft.com/azure/aks/create-node-pools",
    "user_node_pool.min_count": "https://learn.microsoft.com/azure/aks/cluster-autoscaler-overview",
    "user_node_pool.max_count": "https://learn.microsoft.com/azure/aks/cluster-autoscaler-overview",
    "network_policy": "https://learn.microsoft.com/azure/aks/use-network-policies",
    "service_cidr": "https://learn.microsoft.com/azure/aks/configure-azure-cni",
    "dns_service_ip": "https://learn.microsoft.com/azure/aks/configure-azure-cni",
    "pod_cidr": "https://learn.microsoft.com/azure/aks/azure-cni-overlay",
    "automatic_upgrade_channel": "https://learn.microsoft.com/azure/aks/auto-upgrade-cluster",
    "node_os_upgrade_channel": "https://learn.microsoft.com/azure/aks/auto-upgrade-node-os-image",
    "app_gateway_sku": "https://learn.microsoft.com/azure/web-application-firewall/ag/ag-overview",
    "app_gateway_min_capacity": "https://learn.microsoft.com/azure/application-gateway/application-gateway-autoscaling-zone-redundant",
    "app_gateway_max_capacity": "https://learn.microsoft.com/azure/application-gateway/application-gateway-autoscaling-zone-redundant",
    "log_analytics_retention_days": "https://learn.microsoft.com/azure/azure-monitor/logs/data-retention-configure",
}


def set_reference(ws, row, col, setting):
    """Put a clickable Microsoft Learn link in the Reference column."""
    cell = ws.cell(row=row, column=col)
    url = REFERENCES.get(setting)
    if url:
        cell.value = "Microsoft Learn \u2197"
        cell.hyperlink = url
        cell.font = link_font
        cell.alignment = wrap_alignment


# ============================================================================
# TAB 1 — Bootstrap Decisions
# ============================================================================
ws1 = wb.active
ws1.title = "Bootstrap Decisions"

ws1.merge_cells("A1:H1")
ws1["A1"].value = "AKS Application Landing Zone — Planning Checklist"
ws1["A1"].font = title_font
ws1["A1"].alignment = Alignment(vertical="center")
ws1.row_dimensions[1].height = 30

ws1.merge_cells("A2:H2")
ws1["A2"].value = (
    "Fill in the yellow columns. Each row matches one question the wizard will ask you. "
    "Use 'Reference' for the Microsoft Learn docs and 'Comments' for your own notes."
)
ws1["A2"].font = subtitle_font
ws1.row_dimensions[2].height = 20

for col, h in enumerate(["#", "Setting", "What it is", "Example / options", "Default", "Your value", "Reference", "Comments"], 1):
    ws1.cell(row=4, column=col, value=h)
style_header(ws1, 4, 8)

ws1.column_dimensions["A"].width = 6
ws1.column_dimensions["B"].width = 38
ws1.column_dimensions["C"].width = 55
ws1.column_dimensions["D"].width = 55
ws1.column_dimensions["E"].width = 22
ws1.column_dimensions["F"].width = 28
ws1.column_dimensions["G"].width = 20
ws1.column_dimensions["H"].width = 34

# (number, setting, what it is, example/options, default)
decisions = [
    ("", "SCENARIO", "", "", ""),
    ("0a", "scenario",
        "Pick a starting point. Each one chooses sensible defaults for cluster size, security, and add-ons. You can still change anything afterwards.",
        "single_region_baseline  (standard, one region)\n"
        "multi_region_baseline   (two regions, GitOps)\n"
        "single_region_regulated (PCI / FIPS / Istio)\n"
        "multi_region_regulated  (PCI, two regions)",
        "single_region_baseline"),
    ("0b", "secondary_location",
        "Second Azure region (only for multi-region scenarios). Leave blank otherwise.",
        "westeurope, northeurope, eastus2 ...",
        ""),
    ("0c", "topology",
        "How the AKS landing zone connects to the network. 'spoke' peers to an existing ALZ hub and routes egress through the hub firewall (needs Decisions 3 & 4). 'hub_and_spoke' is greenfield: this run also creates a new hub VNet (+ optional Azure Firewall) for you. 'standalone' has no hub, no peering, and uses a NAT gateway for egress (Decisions 3 and 4 are ignored).",
        "spoke         (peer to existing ALZ hub)\n"
        "hub_and_spoke (create a new hub this run)\n"
        "standalone    (no hub, NAT egress only)",
        "spoke"),
    ("0d", "global_lb_type",
        "Global load balancer for multi-region scenarios only. Spreads traffic across both regions and fails over automatically. Ignored for single-region.",
        "front_door      (Azure Front Door Premium; anycast HTTP/S, WAF)\n"
        "traffic_manager (DNS-based, priority failover)\n"
        "none            (single region)",
        "front_door (multi-region) / none (single)"),

    ("", "WHERE TO DEPLOY", "", "", ""),
    ("1", "bootstrap_location",
        "Main Azure region for the cluster and supporting resources.",
        "swedencentral, westeurope, eastus2 ...",
        "swedencentral"),
    ("2", "aks_landing_zone_subscription_id",
        "Subscription where the AKS cluster will be created.",
        "Subscription ID (GUID).\nLeave blank to use whatever `az` is logged into.",
        "(current az subscription)"),
    ("3", "connectivity_subscription_id",
        "Subscription that holds the hub network (existing for 'spoke', or new for 'hub_and_spoke'). Required for spoke and hub_and_spoke; leave blank for standalone.",
        "Subscription ID (GUID).",
        ""),

    ("", "HUB NETWORK — EXISTING HUB (only when topology = spoke)", "", "", ""),
    ("4a", "hub_vnet_resource_id",
        "Full ID of your hub VNet. The wizard lists hubs found in Decision 3 and fills this in. Leave blank when topology = standalone or hub_and_spoke.",
        "/subscriptions/<id>/resourceGroups/<rg>/providers/Microsoft.Network/virtualNetworks/<name>",
        ""),
    ("4b", "hub_firewall_private_ip",
        "Private IP of your hub firewall. Traffic leaving the cluster routes through this. Leave blank when topology = standalone or hub_and_spoke.",
        "10.0.0.4",
        "10.0.0.4"),

    ("", "HUB NETWORK — NEW HUB (only when topology = hub_and_spoke)", "", "", ""),
    ("4c", "hub_vnet_address_space",
        "Address space for the NEW hub VNet this run creates. Must not overlap with the spoke or anything you peer to.",
        "10.0.0.0/16",
        "10.0.0.0/16"),
    ("4d", "hub_firewall_subnet_address_prefix",
        "AzureFirewallSubnet prefix (must be /26 or larger, inside the hub address space).",
        "10.0.0.0/26",
        "10.0.0.0/26"),
    ("4e", "hub_deploy_firewall",
        "Deploy an Azure Firewall in the new hub? If no, you get the hub VNet + AzureFirewallSubnet only and can attach a firewall later.",
        "true | false",
        "true"),
    ("4f", "hub_firewall_sku_tier",
        "Azure Firewall tier (only when 4e is true). Premium adds TLS inspection, IDPS, and URL filtering.",
        "Standard | Premium",
        "Standard"),

    ("", "SPOKE NETWORK (the new VNet for AKS)", "", "", ""),
    ("5a", "spoke_vnet_address_space",
        "Address range for the new spoke VNet. Must not overlap with any other VNet.",
        "10.10.0.0/16",
        "10.10.0.0/16"),
    ("5b", "subnet_address_prefix_aks_system_nodes",
        "Small subnet for the system node pool (runs cluster add-ons only).",
        "10.10.0.0/24  (256 addresses)",
        "10.10.0.0/24"),
    ("5c", "subnet_address_prefix_aks_user_nodes",
        "Bigger subnet for your application workloads.",
        "10.10.16.0/22  (1024 addresses)",
        "10.10.16.0/22"),
    ("5d", "subnet_address_prefix_aks_api_server",
        "Subnet for the private Kubernetes API server.",
        "10.10.20.0/28  (Azure requires at least /28)",
        "10.10.20.0/28"),
    ("5e", "subnet_address_prefix_app_gateway",
        "Subnet for Application Gateway (web traffic in).",
        "10.10.21.0/24",
        "10.10.21.0/24"),
    ("5f", "subnet_address_prefix_private_endpoints",
        "Subnet for private endpoints (ACR, Key Vault, storage).",
        "10.10.22.0/24",
        "10.10.22.0/24"),
    ("5g", "subnet_address_prefix_ingress",
        "Subnet for the internal load balancer that fronts your apps.",
        "10.10.23.0/24",
        "10.10.23.0/24"),
    ("5h", "subnet_address_prefix_agc",
        "Subnet for Application Gateway for Containers (ALB). Only used when enable_agc (11p) is true. Delegated, minimum /24.",
        "10.10.24.0/24",
        "10.10.24.0/24"),

    ("", "CLUSTER SETTINGS", "", "", ""),
    ("6a", "kubernetes_version",
        "Kubernetes version. The wizard lists the latest versions available in your region — usually just pick the top one.",
        "Example: 1.33.6",
        "(latest in region)"),
    ("6b", "aks_sku_tier",
        "Cluster pricing tier. Standard is fine for most clusters; Premium adds long-term support and is required for regulated scenarios.",
        "Free | Standard | Premium",
        "Standard"),
    ("6c", "aks_private_cluster",
        "Hide the Kubernetes API behind your private network (no public endpoint).",
        "true  (recommended)\nfalse (only for internet-facing demos)",
        "true"),
    ("6d", "aks_admin_group_object_ids",
        "Entra ID group(s) whose members can manage the cluster. Create the group in Entra ID first and paste the object ID.",
        '["00000000-0000-0000-0000-000000000000"]',
        "[]"),

    ("", "WHERE STATE LIVES", "", "", ""),
    ("7", "bootstrap_subscription_id",
        "Subscription that holds the Terraform state storage and the runner. Usually the same as Decision 2.",
        "Subscription ID (GUID), or leave blank.",
        "(same as Decision 2)"),

    ("", "NAMING", "", "", ""),
    ("8a", "service_name",
        "Short name of your workload. Used in every resource name.",
        "aksapplz, payments, shop",
        "aksapplz"),
    ("8b", "environment_name",
        "Environment label. Used in every resource name.",
        "prod, dev, test",
        "prod"),
    ("8c", "postfix_number",
        "Number that makes resource names unique. Bump this if you redeploy.",
        "1, 2, 3 ...",
        "1"),

    ("", "RUNNERS", "", "", ""),
    ("9a", "use_self_hosted_runners",
        "Run GitHub Actions on a small container inside Azure (needed for private clusters so CI/CD can reach the API).",
        "true  (required for private clusters)\nfalse (use GitHub-hosted runners)",
        "true"),
    ("9b", "use_private_networking",
        "Put the runner and Container Registry on a private network. Recommended.",
        "true  (recommended)\nfalse (public networking, ACR Basic)",
        "true"),

    ("", "GITHUB", "", "", ""),
    ("10a", "github_personal_access_token",
        "GitHub token used to create the repos. You will be prompted (input is hidden); you do NOT need to put it here.",
        "Scopes needed: repo, workflow, admin:org (Read & Write members).",
        "(prompted in wizard)"),
    ("10b", "github_runners_personal_access_token",
        "Second token used to register the self-hosted runner. Also prompted; only asked when 9a is true.",
        "Scope needed: admin:org (full).",
        "(prompted in wizard)"),
    ("10c", "github_organization_name",
        "GitHub organisation name where the two repos will be created.",
        "contoso",
        ""),
    ("10d", "apply_approvers",
        "GitHub usernames who must approve before a deployment goes live.",
        '["alice", "bob"]',
        "[]"),

    ("", "ADD-ONS (turn things on or off)", "", "", ""),
    ("11a", "enable_defender",
        "Scans the cluster for security threats in real time.",
        "true | false",
        "true"),
    ("11b", "enable_workload_identity",
        "Lets pods talk to Azure services without passwords.",
        "true | false",
        "true"),
    ("11c", "enable_azure_policy",
        "Enforces rules in the cluster (e.g. block containers running as root).",
        "true | false",
        "true"),
    ("11d", "enable_prometheus",
        "Collects cluster metrics.",
        "true | false",
        "true"),
    ("11e", "enable_grafana",
        "Dashboards for the metrics above.",
        "true | false",
        "true"),
    ("11f", "enable_app_gateway",
        "Application Gateway with a Web Application Firewall in front of the cluster.",
        "true | false",
        "true"),
    ("11p", "enable_agc",
        "Application Gateway for Containers (ALB). Provisions a delegated subnet (5h) + NSG; the in-cluster ALB Controller manages the gateway. Coexists with enable_app_gateway.",
        "true | false",
        "false"),
    ("11g", "enable_keda",
        "Auto-scales pods based on events (queue length, HTTP requests, etc.).",
        "true | false",
        "true"),
    ("11h", "enable_vpa",
        "Auto-adjusts CPU and memory requests for pods.",
        "true | false",
        "false  (true for regulated / multi-region)"),
    ("11i", "enable_node_auto_provisioning",
        "Cluster creates new node sizes on demand instead of using fixed pools.",
        "true | false",
        "false  (true for multi-region)"),
    ("11j", "enable_istio",
        "Service mesh with mTLS between pods. Needed for some compliance standards.",
        "true | false",
        "false  (true for regulated)"),
    ("11k", "enable_flux",
        "GitOps — the cluster pulls its configuration from a Git repo.",
        "true | false",
        "false  (true for multi-region)"),
    ("11l", "enable_dapr",
        "Building blocks for microservices (pub/sub, state, secrets).",
        "true | false",
        "false"),
    ("11m", "enable_fips",
        "Use FIPS 140-2 compliant node OS (US Federal / PCI).",
        "true | false",
        "false  (true for regulated)"),
    ("11n", "enable_backup",
        "Azure Backup for cluster resources and persistent volumes.",
        "true | false",
        "false  (true for regulated / multi-region)"),
    ("11o", "enable_cost_analysis",
        "Cost breakdown per namespace in the Azure portal.",
        "true | false",
        "false  (true for regulated)"),
]

row = 5
for num, setting, what, opts, default in decisions:
    if num == "":
        ws1.cell(row=row, column=1, value="")
        ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws1.cell(row=row, column=2, value=setting)
        style_section(ws1, row, 8)
    else:
        ws1.cell(row=row, column=1, value=num)
        ws1.cell(row=row, column=2, value=setting)
        ws1.cell(row=row, column=3, value=what)
        ws1.cell(row=row, column=4, value=opts)
        ws1.cell(row=row, column=5, value=default)
        ws1.cell(row=row, column=6, value="")
        style_data(ws1, row, 8, input_col=6)
        set_reference(ws1, row, 7, setting)
        ws1.cell(row=row, column=8).fill = input_fill
        ws1.cell(row=row, column=1).alignment = center_alignment
    row += 1

bool_settings = {
    "aks_private_cluster", "use_self_hosted_runners", "use_private_networking",
    "enable_defender", "enable_workload_identity", "enable_azure_policy",
    "enable_prometheus", "enable_grafana", "enable_app_gateway", "enable_keda",
    "enable_vpa", "enable_node_auto_provisioning", "enable_istio",
    "enable_flux", "enable_dapr", "enable_fips", "enable_backup",
    "enable_cost_analysis", "enable_agc", "hub_deploy_firewall",
}
for r in range(5, row):
    s = ws1.cell(row=r, column=2).value
    if s == "scenario":
        dropdown(ws1, r, 6, ["single_region_baseline", "multi_region_baseline",
                              "single_region_regulated", "multi_region_regulated"])
    elif s == "topology":
        dropdown(ws1, r, 6, ["spoke", "hub_and_spoke", "standalone"])
    elif s == "global_lb_type":
        dropdown(ws1, r, 6, ["front_door", "traffic_manager", "none"])
    elif s == "hub_firewall_sku_tier":
        dropdown(ws1, r, 6, ["Standard", "Premium"])
    elif s == "aks_sku_tier":
        dropdown(ws1, r, 6, ["Free", "Standard", "Premium"])
    elif s in bool_settings:
        dropdown(ws1, r, 6, ["true", "false"])


# ============================================================================
# TAB 2 — Advanced Cluster Settings
# ============================================================================
ws2 = wb.create_sheet("Advanced Cluster Settings")

ws2.merge_cells("A1:G1")
ws2["A1"].value = "Advanced Cluster Settings"
ws2["A1"].font = title_font
ws2["A1"].alignment = Alignment(vertical="center")
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:G2")
ws2["A2"].value = (
    "These come from the scenario you picked (see Tab 1, Decision 0a). "
    "Only change them if you have a specific reason. Edit `aks-landing-zone.auto.tfvars` "
    "in the generated repo to override."
)
ws2["A2"].font = subtitle_font
ws2.row_dimensions[2].height = 30

for col, h in enumerate(["Setting", "What it does", "Example / options", "Default", "Your value", "Reference", "Comments"], 1):
    ws2.cell(row=4, column=col, value=h)
style_header(ws2, 4, 7)

ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 55
ws2.column_dimensions["C"].width = 45
ws2.column_dimensions["D"].width = 22
ws2.column_dimensions["E"].width = 28
ws2.column_dimensions["F"].width = 20
ws2.column_dimensions["G"].width = 34

advanced = [
    ("", "SYSTEM NODE POOL  (small pool that runs Kubernetes itself)", "", "", ""),
    ("system_node_pool.vm_size",
        "Size of each system node.",
        "Standard_D4ds_v5 (4 vCPU)\nStandard_D8ds_v5 (8 vCPU)",
        "Standard_D4ds_v5"),
    ("system_node_pool.min_count",
        "Smallest number of system nodes.",
        "2 keeps the cluster healthy if one node fails.",
        "2"),
    ("system_node_pool.max_count",
        "Largest number of system nodes.",
        "3–5 is plenty for system pods.",
        "5"),

    ("", "USER NODE POOL  (where your apps actually run)", "", "", ""),
    ("user_node_pool.vm_size",
        "Size of each application node.",
        "Standard_D4ds_v5, Standard_D8ds_v5, Standard_D16ds_v5",
        "Standard_D4ds_v5"),
    ("user_node_pool.min_count",
        "Smallest number of application nodes.",
        "2 or more for production.",
        "2"),
    ("user_node_pool.max_count",
        "Largest number of application nodes (autoscaler limit).",
        "Size to your peak traffic.",
        "20"),

    ("", "NETWORKING", "", "", ""),
    ("network_policy",
        "Which pod-to-pod firewall to use inside the cluster.",
        "calico  (baseline)\nazure   (regulated; NPM)",
        "calico (azure for regulated)"),
    ("service_cidr",
        "IP range Kubernetes uses for its internal services. Must not overlap any VNet.",
        "172.16.0.0/16",
        "172.16.0.0/16"),
    ("dns_service_ip",
        "IP of the cluster DNS. Must sit inside service_cidr.",
        "172.16.0.10",
        "172.16.0.10"),
    ("pod_cidr",
        "IP range for pods (used by Azure CNI Overlay).",
        "192.168.0.0/16",
        "192.168.0.0/16"),

    ("", "UPGRADES", "", "", ""),
    ("automatic_upgrade_channel",
        "When AKS should auto-install Kubernetes patches.",
        "patch | stable | rapid | node-image | none",
        "patch"),
    ("node_os_upgrade_channel",
        "When AKS should auto-patch the node OS.",
        "NodeImage | SecurityPatch | None",
        "NodeImage"),

    ("", "APPLICATION GATEWAY (web traffic + WAF)", "", "", ""),
    ("app_gateway_sku",
        "Application Gateway tier. WAF_v2 includes the Web Application Firewall.",
        "WAF_v2 | Standard_v2",
        "WAF_v2"),
    ("app_gateway_min_capacity",
        "Always-on capacity.",
        "1",
        "1"),
    ("app_gateway_max_capacity",
        "Maximum capacity under load.",
        "Pick based on expected traffic.",
        "10"),

    ("", "LOGS & DASHBOARDS", "", "", ""),
    ("log_analytics_retention_days",
        "How long to keep cluster logs. Longer = more cost.",
        "30, 90, 365 ...",
        "30"),
]

row2 = 5
for tup in advanced:
    if tup[0] == "":
        # section header row uses column B for label
        ws2.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=7)
        ws2.cell(row=row2, column=1, value=tup[1])
        style_section(ws2, row2, 7)
    else:
        setting, what, example, default = tup
        ws2.cell(row=row2, column=1, value=setting)
        ws2.cell(row=row2, column=2, value=what)
        ws2.cell(row=row2, column=3, value=example)
        ws2.cell(row=row2, column=4, value=default)
        ws2.cell(row=row2, column=5, value="")
        style_data(ws2, row2, 7, input_col=5)
        set_reference(ws2, row2, 6, setting)
        ws2.cell(row=row2, column=7).fill = input_fill
    row2 += 1

for r in range(5, row2):
    s = ws2.cell(row=r, column=1).value
    if s == "network_policy":
        dropdown(ws2, r, 5, ["calico", "azure", "cilium"])
    elif s == "app_gateway_sku":
        dropdown(ws2, r, 5, ["WAF_v2", "Standard_v2"])
    elif s == "automatic_upgrade_channel":
        dropdown(ws2, r, 5, ["patch", "stable", "rapid", "node-image", "none"])
    elif s == "node_os_upgrade_channel":
        dropdown(ws2, r, 5, ["NodeImage", "SecurityPatch", "None"])


# ============================================================================
# TAB 3 — How to use this workbook
# ============================================================================
ws3 = wb.create_sheet("How to use")
ws3.column_dimensions["A"].width = 110

intro = [
    ("AKS Application Landing Zone — Planning Checklist", title_font),
    ("", normal_font),
    ("This workbook is the planning sheet for `Deploy-AKSLandingZone`.", normal_font),
    ("Fill it in with your team before you run the wizard, then keep it for the record.", normal_font),
    ("", normal_font),

    ("How to use it", section_font),
    ("", normal_font),
    ("1. Open the 'Bootstrap Decisions' tab and fill in every yellow cell.", normal_font),
    ("   - Pick a scenario first (row 0a) — it sets sensible defaults for the rest.", normal_font),
    ("   - Use the 'Reference' column links to read the Microsoft Learn docs for any setting.", normal_font),
    ("   - Use the 'Comments' column (yellow) to record why you chose a value — handy for your team and auditors.", normal_font),
    ("   - Pick a topology (row 0c) — 'spoke' if you already have an ALZ hub, 'hub_and_spoke' to create a new hub this run, or 'standalone' for an isolated subscription. Standalone skips Decisions 3 and 4.", normal_font),
    ("   - Don't put GitHub tokens in the workbook. The wizard asks for them with hidden input.", normal_font),
    ("", normal_font),
    ("2. Open 'Advanced Cluster Settings' only if you need to change cluster sizing, networking,", normal_font),
    ("   or upgrade behaviour. Most teams leave this tab alone.", normal_font),
    ("", normal_font),
    ("3. Run the wizard and use the workbook as your reference:", normal_font),
    ("       Import-Module .\\ALZ.AKS\\ALZ.AKS.psd1 -Force", normal_font),
    ("       Deploy-AKSLandingZone", normal_font),
    ("", normal_font),

    ("What the wizard does", section_font),
    ("", normal_font),
    ("Phase A — asks you the questions in this workbook and writes config files.", normal_font),
    ("Phase B — creates the Azure resources, GitHub repos, runner, and pushes the Terraform code.", normal_font),
    ("Phase C — your PRs in the new GitHub repo run plan/apply and deploy the cluster.", normal_font),
    ("", normal_font),

    ("GitHub tokens — scopes you need", section_font),
    ("", normal_font),
    ("Main token (always):    repo, workflow, admin:org (Members Read & Write)", normal_font),
    ("Runner token (if 9a is true):  admin:org (full)", normal_font),
    ("", normal_font),

    ("Resource names you will see", section_font),
    ("", normal_font),
    ("Pattern: {service}-{env}-{region-short}-{number}", normal_font),
    ("Example with defaults (aksapplz, prod, swedencentral, 1):", normal_font),
    ("  Resource group for state:     rg-aksapplz-prod-sc-001", normal_font),
    ("  Resource group for identity:  rg-aksapplz-prod-sc-identity", normal_font),
    ("  Resource group for runner:    rg-aksapplz-prod-sc-agents     (only with self-hosted runners)", normal_font),
    ("  GitHub repository:            aksapplz-prod", normal_font),
    ("  GitHub templates repo:        aksapplz-prod-templates", normal_font),
    ("  Approver team:                aksapplz-prod-approvers", normal_font),
]

for i, (text, font) in enumerate(intro, 1):
    cell = ws3.cell(row=i, column=1, value=text)
    cell.font = font
    cell.alignment = Alignment(wrap_text=True, vertical="top")

wb.move_sheet("How to use", offset=-2)

# ─── Save ──────────────────────────────────────────────────────────────────
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "checklist.xlsx")
wb.save(output_path)
print(f"Checklist saved to: {output_path}")
